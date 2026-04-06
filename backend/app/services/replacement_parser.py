"""
Replacement file parser — processes rail replacement records from Excel files.

FILE PURPOSE:
    This file handles the importing of rail replacement records. When a
    section of rail is physically replaced in the field, the event is recorded
    in an Excel file (called an "RMA" file — Rail Maintenance Activity).

    The system needs to know about replacements because they "reset" the wear
    cycle. After a replacement, wear starts from zero again, so predictions
    should only use measurements taken AFTER the replacement.

    This file handles TWO types of input files:

    1. ALREADY-STANDARDIZED FILES:
       These have a sheet called "Standardized_Output" with clean, consistent
       columns. They can be read directly.

    2. RAW RMA FILES:
       These are the original maintenance records, which have a different
       layout. The parser applies the same logic as the legacy VBA macro to
       convert them into the standardized format:
       - Column B = Plan TOA Date Start
       - Column C = Plan TOA Date End (this is the date we use)
       - Column D = Location description
       - Column E = Bound (SBL, SBR, NBL, NBR, or plain SB/NB)
       - Column G = Chainage From (e.g. "39+772.000")
       - Column H = Chainage To
       Data starts at row 6 (first 5 rows are headers).

    When a replacement is processed, the system creates a 0.0mm wear
    measurement at the replacement date for every chainage in the replaced
    range. This marks the "fresh start" in the measurement history.

    Bound codes include rail side information:
      SBL = Southbound Left rail
      SBR = Southbound Right rail
      NBL = Northbound Left rail
      NBR = Northbound Right rail
      SB  = Southbound (both rails)
      NB  = Northbound (both rails)
"""

from __future__ import annotations

import io
import math
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import List, Optional

import openpyxl
from sqlalchemy.orm import Session

from app.models.track import Chainage
from app.models.measurement import WearMeasurement
from app.models.upload import UploadLog


@dataclass
class ReplacementEntry:
    """
    A single standardized replacement record.

    Represents one row of replacement data after it has been normalised
    into a consistent format (regardless of whether it came from a
    standardized file or a raw RMA file).
    """
    location: str               # Description of where the replacement happened
    chainage_from: float        # Start of the replaced range (e.g. 39772.0)
    chainage_to: float          # End of the replaced range (e.g. 39850.0)
    replacement_date: date      # When the replacement was done
    bound: str                  # "SB" or "NB"
    rail_location: str          # "L" (left), "R" (right), or "NIL" (both/unknown)


@dataclass
class ReplacementParseResult:
    """
    Container for the results of processing a replacement file.

    Tracks how many entries were processed, how many measurements were
    created, and what was skipped or errored.
    """
    total_rows: int = 0                    # Total data rows in the file
    valid_entries: int = 0                 # Rows that were successfully processed
    measurements_created: int = 0          # Number of 0mm measurements created
    skipped_no_date: int = 0              # Rows skipped because no date was found
    skipped_bad_bound: int = 0            # Rows skipped because bound was invalid
    skipped_bad_chainage: int = 0         # Rows skipped because chainage was invalid
    skipped_no_match: int = 0             # Rows skipped because no matching chainages found
    affected_chainage_ids: List[int] = field(default_factory=list)  # Chainages that were affected
    entries: List[ReplacementEntry] = field(default_factory=list)   # Parsed entries
    errors: List[str] = field(default_factory=list)                 # Error messages


def _parse_chainage(val) -> Optional[float]:
    """
    Parse a chainage value from the replacement file.

    Chainage values can appear in two formats:
    1. Plain number: 39772 or 39772.000
    2. Plus-sign format: "39+772.000" (which means 39772.000)

    The plus-sign format is commonly used in railway engineering to separate
    the kilometre portion from the metre portion.

    Args:
        val: The raw chainage value from the file.

    Returns:
        A float chainage value, or None if it can't be parsed.
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ('', 'none', 'nan'):
        return None
    # Remove the "+" separator (e.g. "39+772.000" becomes "39772.000")
    s = s.replace('+', '')
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_date(val) -> Optional[date]:
    """
    Parse a date from various formats commonly found in replacement files.

    Handles:
    - Python datetime objects (from openpyxl)
    - Python date objects
    - Date strings in various formats (UK, US, ISO, with/without time)

    Args:
        val: The raw date value from the file.

    Returns:
        A date object, or None if the value can't be parsed as a date.
    """
    if val is None:
        return None
    # If openpyxl already parsed it as a datetime, just extract the date part
    if isinstance(val, datetime):
        return val.date()
    # If it's already a date object, return as-is
    if isinstance(val, date):
        return val
    # Try parsing as a string in various formats
    s = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y',
                '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _normalize_bound(raw_bound: str) -> tuple[str, str]:
    """
    Normalize the bound (direction) code and extract the rail side.

    The raw files use codes like "SBL" (Southbound Left) or "NBR"
    (Northbound Right). This function splits them into the direction
    and the rail side.

    Examples:
        "SBL" -> ("SB", "L")  — Southbound, Left rail
        "NBR" -> ("NB", "R")  — Northbound, Right rail
        "SB"  -> ("SB", "NIL") — Southbound, rail not specified
        "NB"  -> ("NB", "NIL") — Northbound, rail not specified

    Args:
        raw_bound: The raw bound code from the file.

    Returns:
        A tuple of (bound, rail_location) e.g. ("SB", "L").
    """
    b = raw_bound.strip().upper()
    mapping = {
        'SBL': ('SB', 'L'), 'SBR': ('SB', 'R'),
        'NBL': ('NB', 'L'), 'NBR': ('NB', 'R'),
        'SB': ('SB', 'NIL'), 'NB': ('NB', 'NIL'),
        'XB': ('XB', 'NIL'), 'BB': ('BB', 'NIL'),
        'XBL': ('XB', 'L'), 'XBR': ('XB', 'R'),
        'BBL': ('BB', 'L'), 'BBR': ('BB', 'R'),
    }
    return mapping.get(b, (b, 'NIL'))


def _standardize_raw_sheet(ws) -> List[list]:
    """
    Apply the VBA macro logic to a raw RMA worksheet, converting it to a
    standardized format.

    The raw RMA files have a specific layout:
    - Rows 1-5 are headers/metadata (skipped)
    - Row 6 onwards is actual data
    - Column C (index 2) = Plan To End Date (the replacement date)
    - Column D (index 3) = Location description
    - Column E (index 4) = Bound (SBL, SBR, NBL, NBR, SB, NB)
    - Column G (index 6) = Chainage From
    - Column H (index 7) = Chainage To

    This function reads those columns and creates a standardized output
    with consistent column names.

    Args:
        ws: An openpyxl worksheet object.

    Returns:
        A list of rows, where the first row is the header and subsequent
        rows are the standardized data.
    """
    rows = list(ws.iter_rows(values_only=True))

    # Start with a header row
    standardized = [['Location', 'Chainage From', 'Chainage To',
                     'Plan To End Date', 'Bound', 'Rail Location']]

    # Data starts at row 6 (index 5) — skip the header rows
    for i in range(5, len(rows)):  # Start at row 6 (index 5)
        row = rows[i]
        # Skip rows that don't have enough columns
        if len(row) < 8:
            continue

        # Read the relevant columns from the raw format
        plan_date = row[2]       # Column C = Plan To End Date (replacement date)
        location = row[3]        # Column D = Location description
        original_bound = row[4]  # Column E = Bound (direction + rail side)
        ch_from_raw = row[6]     # Column G = Chainage From
        ch_to_raw = row[7]       # Column H = Chainage To

        # Skip completely empty rows
        if not any([location, ch_from_raw, ch_to_raw, original_bound]):
            continue

        # Parse chainage values (handles "39+772.000" format)
        ch_from = _parse_chainage(ch_from_raw)
        ch_to = _parse_chainage(ch_to_raw)

        # Normalize the bound code (e.g. "SBL" -> "SB" + "L")
        bound_str = str(original_bound).strip() if original_bound else ''
        bound, rail_loc = _normalize_bound(bound_str)

        # Add the standardized row
        standardized.append([
            str(location).strip() if location else '',
            ch_from, ch_to, plan_date, bound, rail_loc
        ])

    return standardized


def parse_replacement_file(
    file_content: bytes,
    filename: str,
) -> tuple[List[list], str]:
    """
    Parse a replacement Excel file and return standardized rows.

    This function auto-detects the file format:
    1. If the file has a "Standardized_Output" sheet, it reads directly from that
    2. Otherwise, it finds the RMA data sheet and applies the VBA macro logic

    This two-step approach means the system can handle both old-format and
    new-format replacement files.

    Args:
        file_content: Raw bytes of the uploaded Excel file.
        filename: Original filename.

    Returns:
        A tuple of (rows, format_detected) where:
        - rows: list of lists (first row is header, rest is data)
        - format_detected: "standardized" or "raw"
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_content),
        read_only=True, data_only=True, keep_vba=False
    )

    # Check if the file already has a standardized output sheet
    if 'Standardized_Output' in wb.sheetnames:
        ws = wb['Standardized_Output']
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        return rows, "standardized"

    # Otherwise, find the raw RMA data sheet.
    # We skip sheets that are known to be non-data sheets (like crossings,
    # monthly reports, totals, etc.)
    rma_sheet = None
    for sn in wb.sheetnames:
        sn_lower = sn.lower()
        # Skip known non-data sheets
        if any(skip in sn_lower for skip in ['x-ing', 'monthly', 'total', 'sheet1', 'lookup']):
            continue
        # Look for sheets named after RMA or containing year numbers
        if 'rma' in sn_lower or 'rail replacement' in sn_lower or any(str(y) in sn for y in range(2014, 2030)):
            rma_sheet = sn
            break

    # If we can't identify the RMA sheet, fall back to the first sheet
    if rma_sheet is None:
        rma_sheet = wb.sheetnames[0]

    # Apply the VBA macro logic to convert raw format to standardized
    ws = wb[rma_sheet]
    rows = _standardize_raw_sheet(ws)
    wb.close()
    return rows, "raw"


def process_replacement_rows(
    rows: List[list],
    db: Session,
    upload_log: UploadLog,
) -> ReplacementParseResult:
    """
    Process standardized replacement rows and create 0mm wear measurements.

    This is the second stage of replacement processing. After the file has
    been parsed into standardized rows (by parse_replacement_file), this
    function:

    1. For each replacement row:
       a. Validates the date, bound, and chainage range
       b. Finds all chainage records within the replaced range
       c. Creates a 0.0mm wear measurement for each affected chainage
          (this marks the "fresh start" after replacement)

    The 0mm measurements are important because they tell the prediction
    system that wear has been reset to zero at this date.

    Args:
        rows: Standardized rows (first row is the header row).
        db: Database session.
        upload_log: The upload log to link the new measurements to.

    Returns:
        A ReplacementParseResult with counts and any errors.
    """
    result = ReplacementParseResult()
    VALID_BOUNDS = {'SB', 'NB', 'XB', 'BB'}

    # Build a cache of all chainages, keyed by (chainage_id, bound).
    # This avoids hitting the database for every single chainage number.
    all_ch = db.query(Chainage).all()
    ch_cache = {}
    for c in all_ch:
        ch_cache[(c.chainage_id, c.bound)] = c

    # Pre-load existing (chainage, date) pairs for duplicate detection
    existing_pairs = set()
    all_meas = db.query(
        WearMeasurement.chainage_id, WearMeasurement.measurement_date
    ).all()
    for cid, mdate in all_meas:
        existing_pairs.add((cid, mdate))

    # Process each row (skip the header at index 0)
    for i, row in enumerate(rows):
        if i == 0:  # Skip the header row
            continue
        result.total_rows += 1

        if len(row) < 5:
            continue

        # Parse the fields from the standardized row
        ch_from = _parse_chainage(row[1])     # Chainage From
        ch_to = _parse_chainage(row[2])       # Chainage To
        rep_date = _parse_date(row[3])        # Replacement date
        bound = str(row[4]).strip().upper() if row[4] else ''  # Direction

        # --- Validation checks ---

        # This checks that the bound (direction) is valid
        if bound not in VALID_BOUNDS:
            result.skipped_bad_bound += 1
            continue

        # This checks that a valid date was found
        if rep_date is None:
            result.skipped_no_date += 1
            continue

        # This checks that both chainage values are valid numbers
        if ch_from is None or ch_to is None:
            result.skipped_bad_chainage += 1
            continue

        # Sanity check: chainage values should be realistic (> 1000)
        if ch_from < 1000 or ch_to < 1000:
            result.skipped_bad_chainage += 1
            continue

        # Sanity check: the replaced range should not be absurdly large
        # (more than 5km would indicate a data error)
        if abs(ch_to - ch_from) > 5000:
            result.skipped_bad_chainage += 1
            continue

        # --- Create 0mm measurements for each chainage in the replaced range ---
        # Convert to integers and ensure lo <= hi
        lo = int(min(ch_from, ch_to))
        hi = int(max(ch_from, ch_to))

        entry_created = 0
        # Loop through every whole-number chainage in the replaced range
        for ch_int in range(lo, hi + 1):
            ch_id = str(ch_int)
            # Look up this chainage in our cache
            chainage = ch_cache.get((ch_id, bound))
            if chainage is None:
                # This chainage doesn't exist in the database — skip it
                continue

            # Skip if we already have a measurement for this chainage+date
            pair_key = (chainage.id, rep_date)
            if pair_key in existing_pairs:
                continue
            existing_pairs.add(pair_key)

            # Create a 0.0mm measurement to mark the replacement.
            # All wear positions are set to 0.0 because the rail is brand new.
            # DTL (XB/BB) uses 5 positions; NEL (NB/SB) uses 2 positions.
            meas = WearMeasurement(
                chainage_id=chainage.id,
                measurement_date=rep_date,
                wear_mm=0.0,
                left_wear_0=0.0, left_wear_90=0.0,
                right_wear_0=0.0, right_wear_90=0.0,
                left_wear_22_5=0.0, left_wear_45=0.0, left_wear_67_5=0.0,
                right_wear_22_5=0.0, right_wear_45=0.0, right_wear_67_5=0.0,
                source_file=f"replacement_upload",
                upload_id=upload_log.id,
            )
            db.add(meas)
            entry_created += 1

            # Track which chainages were affected (for prediction refitting)
            if chainage.id not in result.affected_chainage_ids:
                result.affected_chainage_ids.append(chainage.id)

        # Update counts based on whether any measurements were created
        if entry_created > 0:
            result.valid_entries += 1
            result.measurements_created += entry_created
        else:
            # No matching chainages found in the database for this range
            result.skipped_no_match += 1

    return result
