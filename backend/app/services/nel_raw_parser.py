"""
NEL raw equipment file parser — converts Excel files from the track inspection
equipment into database records.

FILE PURPOSE:
    This file reads the Excel files that come directly from the NEL (North
    East Line) track inspection equipment. These files have a specific fixed
    column layout that is different from standard CSV uploads.

    The equipment records rail wear measurements at every metre along the
    track. Each row in the file represents one measurement point (chainage).
    The columns contain:
    - Column A: chainage number (distance marker, e.g. 42915)
    - Columns G-J: wear values at four angular positions
      (left 90 degrees, right 90 degrees, left 0 degrees, right 0 degrees)

    There is also an "extended" format that includes additional columns:
    - Column K: Radius (curve radius in metres; 1000000000 = straight track)
    - Column L: TrackType ("Standard Straight", "Standard Curve", "Premium Curve")

    WHAT THIS PARSER DOES:
    1. Opens the Excel file and reads all rows
    2. Auto-detects whether the train was going northbound or southbound
       (by checking if chainage numbers increase or decrease)
    3. Filters to whole-number chainages only (the equipment records at
       fractional chainages too, but we only keep the whole numbers)
    4. Looks up which inter-station sector each chainage belongs to
    5. Creates new Chainage records in the database if they don't exist yet
    6. Creates WearMeasurement records for each valid data row
    7. (Extended format only) Extracts and saves category/curve information

    This is a Python equivalent of the legacy VBA macro that was previously
    used to process these files in Excel.
"""

from __future__ import annotations

import io
import math
from dataclasses import dataclass, field
from datetime import date

import openpyxl
from sqlalchemy.orm import Session

from app.models.track import Chainage, Track
from app.models.measurement import WearMeasurement
from app.services.nel_sector_lookup import (
    detect_bound, get_sector, is_whole_number, validate_chainage_range,
)

# Mapping from NEL 3-letter station codes to full station names.
# Used when creating new Chainage records to fill in the station name fields.
STATION_CODE_TO_NAME: dict[str, str] = {
    "HBF": "HarbourFront", "OTP": "Outram Park", "CNT": "Chinatown",
    "CQY": "Clarke Quay", "DBG": "Dhoby Ghaut", "LTI": "Little India",
    "FRP": "Farrer Park", "BNK": "Boon Keng", "PTP": "Potong Pasir",
    "WLH": "Woodleigh", "SER": "Serangoon", "KVN": "Kovan",
    "HGN": "Hougang", "BGK": "Buangkok", "SKG": "Sengkang", "PGL": "Punggol",
}


def _sector_to_stations(sector: str) -> tuple[str | None, str | None]:
    """
    Convert a sector code (e.g. "PGL-SKG") into full station names
    (e.g. "Punggol", "Sengkang").

    This is used when creating new Chainage records — we want to store the
    full station names rather than just the codes.

    If the sector code is a single station (e.g. "SKG"), the same station
    name is returned for both start and end.

    Args:
        sector: Sector string like "PGL-SKG" or station code like "SKG".

    Returns:
        A tuple of (start_station_name, end_station_name), or (None, None)
        if the sector code is not recognised.
    """
    # This handles the case where no sector was found
    if not sector or sector == "NULL":
        return None, None

    parts = sector.split("-")

    # This handles inter-station sectors like "PGL-SKG"
    if len(parts) == 2:
        return STATION_CODE_TO_NAME.get(parts[0]), STATION_CODE_TO_NAME.get(parts[1])

    # This handles station areas like "SKG" (within a single station)
    if len(parts) == 1:
        # Same station for both start and end
        name = STATION_CODE_TO_NAME.get(parts[0])
        return name, name

    return None, None


@dataclass
class NELRawParseResult:
    """
    A container that holds the results of parsing an NEL raw equipment file.

    This gives a complete summary of what happened during parsing:
    - How many rows were in the file
    - How many had whole-number chainages (vs fractional interpolation points)
    - How many had actual wear data
    - How many were successfully processed vs skipped/errored
    - Which direction (NB/SB) was detected
    - Which sectors were found
    - How many new chainage records were created
    """
    total_raw_rows: int = 0              # Total data rows in the file
    whole_number_rows: int = 0           # Rows with whole-number chainages (not interpolated)
    rows_with_wear: int = 0             # Rows that had at least one wear value
    rows_accepted: int = 0              # Rows successfully converted to measurements
    rows_skipped: int = 0               # Rows skipped (duplicate chainage+date)
    rows_errored: int = 0               # Rows that failed processing
    bound_detected: str = "NULL"        # Auto-detected direction: "NB", "SB", or "NULL"
    chainage_range: tuple = (0, 0)      # (min_chainage, max_chainage) found in file
    sectors_found: list[str] = field(default_factory=list)     # List of sectors found
    new_chainages_created: int = 0      # How many new Chainage records were auto-created
    errors: list[dict] = field(default_factory=list)           # Error details
    warnings: list[dict] = field(default_factory=list)         # Warning details
    measurements: list[WearMeasurement] = field(default_factory=list)  # Measurement records to save


def _parse_wear(value) -> float | None:
    """
    Parse a wear cell value from the equipment file to a decimal number.

    Wear values can be:
    - Positive numbers (normal wear)
    - Negative numbers (indicates gauge-face wear direction in some equipment)
    - None or empty (no measurement at this position)
    - Placeholder strings like "---" or "N/A"

    Args:
        value: The raw cell value from the Excel file.

    Returns:
        A float number, or None if the value is empty or unparseable.
    """
    if value is None:
        return None

    # This handles numeric values directly (most common case)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)

    # This handles string values (less common — usually placeholders)
    s = str(value).strip()
    if s in ("", "---", "N/A", "NA", "-"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_nel_raw(
    file_content: bytes,
    filename: str,
    measurement_date: date,
    db: Session,
) -> NELRawParseResult:
    """
    Parse a raw NEL equipment Excel file and create WearMeasurement records.

    This is the main function for processing standard NEL equipment files
    (without category columns). It reads the Excel file, auto-detects the
    travel direction, and creates measurement records for each valid row.

    HOW IT WORKS:
    1. Open the Excel file and read all rows into memory
    2. Collect all chainage values from column A for direction detection
    3. Auto-detect northbound vs southbound from chainage progression
    4. Look up the NEL track in the database
    5. Pre-load existing measurements for duplicate detection
    6. For each row:
       a. Skip non-numeric rows (headers, notes, etc.)
       b. Skip fractional chainages (only keep whole numbers)
       c. Read wear values from columns G, H, I, J
       d. Skip rows where all wear values are empty
       e. Look up the sector for this chainage
       f. Get or create the Chainage record in the database
       g. Skip if this chainage+date already exists (duplicate)
       h. Create a WearMeasurement record

    Args:
        file_content: Raw bytes of the uploaded Excel file.
        filename: Original filename (stored on measurements for traceability).
        measurement_date: The date to assign to all measurements.
        db: Database session for lookups, creation, and duplicate detection.

    Returns:
        An NELRawParseResult with the parsed measurements and metadata.
    """
    result = NELRawParseResult()

    # Step 1: Open the Excel file
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    except Exception as e:
        result.errors.append({"row": 0, "error": f"Cannot read Excel file: {e}"})
        return result

    # Get the active (first) worksheet
    ws = wb.active
    if ws is None:
        result.errors.append({"row": 0, "error": "No active sheet found"})
        wb.close()
        return result

    # Step 2: Read all rows into memory and collect column A values
    # (read_only mode requires iterating through the entire file)
    rows_data: list[list] = []
    column_a_values: list = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # Skip the header row (first row)
        rows_data.append(list(row))
        # Collect column A values for direction detection
        if row[0] is not None:
            column_a_values.append(row[0])

    wb.close()

    result.total_raw_rows = len(rows_data)

    if not rows_data:
        result.errors.append({"row": 0, "error": "File has no data rows"})
        return result

    # Step 3: Detect direction (NB/SB) by checking if chainages go up or down
    result.bound_detected = detect_bound(column_a_values)

    # Step 4: Look up the NEL track record (must already exist in the database)
    nel_track = db.query(Track).filter(Track.track_id == "NEL").first()
    if not nel_track:
        result.errors.append({"row": 0, "error": "NEL track not found in database"})
        return result

    # Step 5: Pre-load existing (chainage, date) pairs for duplicate detection.
    # This avoids trying to insert measurements that already exist.
    existing_pairs: set[tuple[int, date]] = set()
    all_existing = db.query(WearMeasurement.chainage_id, WearMeasurement.measurement_date).all()
    for cid, mdate in all_existing:
        existing_pairs.add((cid, mdate))

    # Cache for chainage lookups — avoids querying the database repeatedly
    # for the same chainage number
    chainage_cache: dict[str, Chainage] = {}
    sectors_seen: set[str] = set()
    min_ch = float("inf")      # Track the smallest chainage found
    max_ch = float("-inf")     # Track the largest chainage found

    # Step 6: Process each data row
    for row_idx, row in enumerate(rows_data):
        row_num = row_idx + 2  # Human-readable row number (1-indexed + header)

        # --- Get the chainage value from column A (first column) ---
        ch_val = row[0] if len(row) > 0 else None

        # Skip non-numeric rows (could be section headers, notes, etc.)
        if not isinstance(ch_val, (int, float)) or isinstance(ch_val, bool):
            continue

        ch_float = float(ch_val)

        # Only process whole-number chainages — skip sub-metre interpolation
        # points (e.g. skip 42915.25, keep 42915.0)
        if not is_whole_number(ch_float):
            continue

        result.whole_number_rows += 1
        ch_int = int(ch_float)

        # Track the min/max chainage range for the result metadata
        if ch_float < min_ch:
            min_ch = ch_float
        if ch_float > max_ch:
            max_ch = ch_float

        # --- Read wear values from fixed column positions ---
        # The NEL equipment file has a fixed layout:
        #   Column G (index 6) = left rail at 90 degrees (gauge face)
        #   Column H (index 7) = right rail at 90 degrees (gauge face)
        #   Column I (index 8) = left rail at 0 degrees (top of rail)
        #   Column J (index 9) = right rail at 0 degrees (top of rail)
        left_90 = _parse_wear(row[6] if len(row) > 6 else None)
        right_90 = _parse_wear(row[7] if len(row) > 7 else None)
        left_0 = _parse_wear(row[8] if len(row) > 8 else None)
        right_0 = _parse_wear(row[9] if len(row) > 9 else None)

        # Skip rows where ALL four wear positions are empty (no data)
        wear_vals = [v for v in [left_0, left_90, right_0, right_90] if v is not None]
        if not wear_vals:
            continue

        result.rows_with_wear += 1

        # --- Look up the inter-station sector for this chainage ---
        sector = get_sector(ch_float, result.bound_detected)
        if sector != "NULL":
            sectors_seen.add(sector)

        # The overall wear is the maximum ABSOLUTE value across all positions.
        # We use absolute value because some equipment reports negative numbers
        # for certain wear directions.
        wear_mm = max(abs(v) for v in wear_vals)

        # --- Get or auto-create the Chainage record ---
        ch_id_str = str(ch_int)
        if ch_id_str not in chainage_cache:
            # Check if this chainage already exists in the database
            existing = db.query(Chainage).filter(
                Chainage.chainage_id == ch_id_str,
                Chainage.bound == result.bound_detected,
            ).first()
            if existing:
                chainage_cache[ch_id_str] = existing
            else:
                # Auto-create a new Chainage record.
                # This happens the first time we see a chainage that isn't in
                # the database yet (e.g. when processing the very first file).
                start_stn, end_stn = _sector_to_stations(sector)
                new_ch = Chainage(
                    chainage_id=ch_id_str,
                    track_id=nel_track.id,
                    bound=result.bound_detected,
                    sector=sector,
                    rail_side="Both",
                    start_station=start_stn,
                    end_station=end_stn,
                )
                db.add(new_ch)
                # flush() saves the new record immediately so it gets an ID,
                # but doesn't commit the transaction yet
                db.flush()
                chainage_cache[ch_id_str] = new_ch
                result.new_chainages_created += 1

        chainage_obj = chainage_cache[ch_id_str]

        # Backfill start/end station names if they were missing from an
        # earlier import (this fills in gaps from older data)
        if chainage_obj.start_station is None and sector and sector != "NULL":
            start_stn, end_stn = _sector_to_stations(sector)
            chainage_obj.start_station = start_stn
            chainage_obj.end_station = end_stn

        # --- Duplicate detection ---
        # Skip if we already have a measurement for this chainage on this date
        if (chainage_obj.id, measurement_date) in existing_pairs:
            result.rows_skipped += 1
            continue

        # --- Create the WearMeasurement record ---
        # NEL equipment only measures at 0 and 90 degrees, so the other
        # three positions (22.5, 45, 67.5) are always set to None.
        meas = WearMeasurement(
            chainage_id=chainage_obj.id,
            measurement_date=measurement_date,
            wear_mm=round(wear_mm, 2),                                    # Round to 2 decimal places
            left_wear_0=round(left_0, 2) if left_0 is not None else None,
            left_wear_90=round(left_90, 2) if left_90 is not None else None,
            right_wear_0=round(right_0, 2) if right_0 is not None else None,
            right_wear_90=round(right_90, 2) if right_90 is not None else None,
            # NEL equipment only measures at 0 and 90 degrees — these are always NULL
            left_wear_22_5=None,
            left_wear_45=None,
            left_wear_67_5=None,
            right_wear_22_5=None,
            right_wear_45=None,
            right_wear_67_5=None,
            source_file=filename,
        )
        result.measurements.append(meas)
        # Add to the duplicate detection set for the current file
        existing_pairs.add((chainage_obj.id, measurement_date))
        result.rows_accepted += 1

    # Set the chainage range in the result metadata
    result.chainage_range = (int(min_ch) if min_ch != float("inf") else 0,
                              int(max_ch) if max_ch != float("-inf") else 0)
    result.sectors_found = sorted(sectors_seen)

    return result


# ---------------------------------------------------------------------------
# Extended format: NEL raw files with Radius and TrackType columns
# ---------------------------------------------------------------------------

# This is the threshold value used by the equipment to indicate "straight track".
# If the Radius column has a value >= 100 million, the track is considered straight.
# (The equipment uses 1,000,000,000 as the sentinel value for straight track.)
STRAIGHT_RADIUS_THRESHOLD = 100_000_000


def _detect_category_columns(header_row: tuple | list) -> tuple[int | None, int | None, int | None]:
    """
    Find the column positions (indices) for Radius, TrackType, and curve
    direction in the header row of an equipment file.

    This scans the header row to find which column numbers contain the
    category-related data. The column positions can vary between files,
    so we detect them dynamically rather than hardcoding.

    Args:
        header_row: The first row of the spreadsheet (column headers).

    Returns:
        A tuple of (radius_column_index, tracktype_column_index,
        direction_column_index). Any of these can be None if not found.
    """
    radius_idx = None
    tracktype_idx = None
    direction_idx = None

    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        # Normalise the header name for matching
        name = str(cell).strip().lower().replace(" ", "")
        if name == "radius":
            radius_idx = i
        elif name == "tracktype":
            tracktype_idx = i
        elif name in ("curvedirection", "curvedir", "left/right/straight", "column13"):
            direction_idx = i

    return radius_idx, tracktype_idx, direction_idx


def _classify_category(radius_val, tracktype_val) -> tuple[int | None, float | None, str]:
    """
    Determine the track category from the Radius and TrackType column values.

    This converts the raw equipment data into our category system:

    Category 1 (Straight Standard):
        - Radius >= 100,000,000 (the equipment uses a very large number for straight)
        - OR TrackType contains "Straight"

    Category 2 (Standard Curve):
        - Radius < 100,000,000 (actual curve radius)
        - AND TrackType is "Standard Curve" (or similar)

    Category 3 (Premium Curve):
        - TrackType contains "Premium"
        - These are tight curves with head-hardened rail

    Args:
        radius_val: Raw value from the Radius column.
        tracktype_val: Raw value from the TrackType column.

    Returns:
        A tuple of (category, curve_radius, rail_type):
        - category: 1, 2, 3, or None if it can't be determined
        - curve_radius: the actual radius in metres, or None for straight track
        - rail_type: "standard" or "premium"
    """
    # --- Parse the radius value ---
    radius_num = None
    if radius_val is not None and isinstance(radius_val, (int, float)) and not isinstance(radius_val, bool):
        fval = float(radius_val)
        # Only accept finite numbers (reject NaN and infinity)
        if math.isfinite(fval):
            radius_num = fval

    # --- Parse the track type string ---
    tt_str = ""
    if tracktype_val is not None:
        s = str(tracktype_val).strip().lower()
        # Skip pandas NaN values which show up as the string "nan"
        if s != "nan":
            tt_str = s

    # --- Classify based on the parsed values ---

    # Check for premium rail first (takes priority over everything else)
    if "premium" in tt_str:
        return 3, radius_num if radius_num and radius_num < STRAIGHT_RADIUS_THRESHOLD else None, "premium"

    # Check if the track is straight
    is_straight = False
    # If the radius is very large (>= 100 million), the equipment considers it straight
    if radius_num is not None and radius_num >= STRAIGHT_RADIUS_THRESHOLD:
        is_straight = True
    # Or if the track type explicitly says "straight"
    elif "straight" in tt_str:
        is_straight = True

    if is_straight:
        return 1, None, "standard"  # Category 1, no curve radius, standard rail

    # Check if it's a standard curve (has a real radius value)
    if radius_num is not None and radius_num < STRAIGHT_RADIUS_THRESHOLD:
        return 2, radius_num, "standard"  # Category 2, actual curve radius

    # If the track type says "curve" but we don't have a valid radius
    if "curve" in tt_str:
        return 2, radius_num, "standard"

    # Can't determine category from the available data
    return None, None, "standard"


@dataclass
class NELRawWithCategoryResult(NELRawParseResult):
    """
    Extended result that includes category information in addition to the
    standard NEL parse results.

    This is used when processing files that have Radius and TrackType columns.
    """
    categories_updated: int = 0                          # How many chainages had their category set
    category_summary: dict[int, int] = field(default_factory=dict)  # Breakdown: {1: 500, 2: 30, 3: 10}


def parse_nel_raw_with_category(
    file_content: bytes,
    filename: str,
    measurement_date: date,
    db: Session,
) -> NELRawWithCategoryResult:
    """
    Parse a raw NEL equipment file that includes Radius and TrackType columns.

    This works the same as parse_nel_raw() for wear measurements, but ALSO
    extracts the track category and curve radius from the additional columns
    and updates the Chainage records accordingly.

    This means a single file upload can both:
    1. Import wear measurement data
    2. Set/update the track category for each chainage

    The function supports both Excel (.xlsx) and CSV (.csv) file formats.

    Args:
        file_content: Raw bytes of the uploaded file.
        filename: Original filename.
        measurement_date: The date to assign to all parsed measurements.
        db: Database session.

    Returns:
        An NELRawWithCategoryResult with measurements, category updates, and metadata.
    """
    result = NELRawWithCategoryResult()

    # --- Read the file (supports both CSV and Excel) ---
    lower = filename.lower()
    rows_data: list[list] = []
    column_a_values: list = []
    radius_col: int | None = None
    tracktype_col: int | None = None

    if lower.endswith((".csv",)):
        # --- CSV file handling ---
        import pandas as pd
        try:
            # Try multiple character encodings (same approach as _read_file)
            for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
                try:
                    df = pd.read_csv(io.BytesIO(file_content), encoding=encoding)
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
            else:
                raise ValueError("Unable to decode CSV")
        except Exception as e:
            result.errors.append({"row": 0, "error": f"Cannot read CSV file: {e}"})
            return result

        # Find the Radius and TrackType columns in the CSV headers
        radius_col, tracktype_col, direction_col = _detect_category_columns(list(df.columns))

        # Convert DataFrame rows to lists (to match the format used by the
        # Excel parser, so the rest of the code works the same way)
        for _, row in df.iterrows():
            vals = row.tolist()
            rows_data.append(vals)
            # Collect column A values for direction detection, skipping NaN
            if vals[0] is not None and not (isinstance(vals[0], float) and pd.isna(vals[0])):
                column_a_values.append(vals[0])
    else:
        # --- Excel file handling ---
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        except Exception as e:
            result.errors.append({"row": 0, "error": f"Cannot read Excel file: {e}"})
            return result

        ws = wb.active
        if ws is None:
            result.errors.append({"row": 0, "error": "No active sheet found"})
            wb.close()
            return result

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                # First row is the header — detect category column positions
                radius_col, tracktype_col, direction_col = _detect_category_columns(row)
                continue
            rows_data.append(list(row))
            if row[0] is not None:
                column_a_values.append(row[0])

        wb.close()

    result.total_raw_rows = len(rows_data)

    if not rows_data:
        result.errors.append({"row": 0, "error": "File has no data rows"})
        return result

    # Detect direction (NB/SB)
    result.bound_detected = detect_bound(column_a_values)

    # Look up the NEL track record
    nel_track = db.query(Track).filter(Track.track_id == "NEL").first()
    if not nel_track:
        result.errors.append({"row": 0, "error": "NEL track not found in database"})
        return result

    # Pre-load existing (chainage, date) pairs for duplicate detection
    existing_pairs: set[tuple[int, date]] = set()
    all_existing = db.query(WearMeasurement.chainage_id, WearMeasurement.measurement_date).all()
    for cid, mdate in all_existing:
        existing_pairs.add((cid, mdate))

    chainage_cache: dict[str, Chainage] = {}
    sectors_seen: set[str] = set()
    min_ch = float("inf")
    max_ch = float("-inf")
    cat_counts: dict[int, int] = {}  # Tracks how many chainages per category

    # --- Process each data row ---
    for row_idx, row in enumerate(rows_data):
        row_num = row_idx + 2  # Human-readable row number

        # Get chainage value from column A
        ch_val = row[0] if len(row) > 0 else None
        # Skip non-numeric rows
        if not isinstance(ch_val, (int, float)) or isinstance(ch_val, bool):
            continue

        ch_float = float(ch_val)
        # Only keep whole-number chainages (skip sub-metre interpolation)
        if not is_whole_number(ch_float):
            continue

        result.whole_number_rows += 1
        ch_int = int(ch_float)

        # Track chainage range
        if ch_float < min_ch:
            min_ch = ch_float
        if ch_float > max_ch:
            max_ch = ch_float

        # Read wear values (same column layout as standard NEL raw)
        left_90 = _parse_wear(row[6] if len(row) > 6 else None)
        right_90 = _parse_wear(row[7] if len(row) > 7 else None)
        left_0 = _parse_wear(row[8] if len(row) > 8 else None)
        right_0 = _parse_wear(row[9] if len(row) > 9 else None)

        # Skip rows where all wear values are empty
        wear_vals = [v for v in [left_0, left_90, right_0, right_90] if v is not None]
        if not wear_vals:
            continue

        result.rows_with_wear += 1

        # Look up the sector for this chainage
        sector = get_sector(ch_float, result.bound_detected)
        if sector != "NULL":
            sectors_seen.add(sector)

        # Calculate overall wear (maximum absolute value)
        wear_mm = max(abs(v) for v in wear_vals)

        # --- Extract category/curvature from the Radius and TrackType columns ---
        # These columns only exist in the "extended" format files.
        radius_val = row[radius_col] if radius_col is not None and len(row) > radius_col else None
        tracktype_val = row[tracktype_col] if tracktype_col is not None and len(row) > tracktype_col else None
        direction_val = row[direction_col] if direction_col is not None and len(row) > direction_col else None

        # Classify the track category from the Radius and TrackType values
        cat, curve_r, rail_t = _classify_category(radius_val, tracktype_val)

        # Parse curve direction (L=left, R=right, S=straight)
        curve_dir = None
        if direction_val is not None:
            d = str(direction_val).strip().lower()
            if d in ("l", "left"):
                curve_dir = "left"
            elif d in ("r", "right"):
                curve_dir = "right"

        # --- Get or create the Chainage record ---
        ch_id_str = str(ch_int)
        if ch_id_str not in chainage_cache:
            existing = db.query(Chainage).filter(
                Chainage.chainage_id == ch_id_str,
                Chainage.bound == result.bound_detected,
            ).first()
            if existing:
                chainage_cache[ch_id_str] = existing
            else:
                # Auto-create a new Chainage record
                start_stn, end_stn = _sector_to_stations(sector)
                new_ch = Chainage(
                    chainage_id=ch_id_str,
                    track_id=nel_track.id,
                    bound=result.bound_detected,
                    sector=sector,
                    rail_side="Both",
                    start_station=start_stn,
                    end_station=end_stn,
                )
                db.add(new_ch)
                db.flush()
                chainage_cache[ch_id_str] = new_ch
                result.new_chainages_created += 1

        chainage_obj = chainage_cache[ch_id_str]

        # Backfill station names if missing
        if chainage_obj.start_station is None and sector and sector != "NULL":
            start_stn, end_stn = _sector_to_stations(sector)
            chainage_obj.start_station = start_stn
            chainage_obj.end_station = end_stn

        # --- Update category/curvature on the chainage record ---
        # This sets the category, curve radius, rail type, and curve direction
        # based on the Radius and TrackType columns in the file.
        if cat is not None:
            chainage_obj.category = cat
            chainage_obj.curve_radius = curve_r
            chainage_obj.rail_type = rail_t
            # Only set curve direction for curved track (categories 2 and 3)
            chainage_obj.curve_direction = curve_dir if cat in (2, 3) else None
            # Count how many chainages per category (for the summary)
            cat_counts[cat] = cat_counts.get(cat, 0) + 1
            result.categories_updated += 1

        # --- Duplicate detection ---
        if (chainage_obj.id, measurement_date) in existing_pairs:
            result.rows_skipped += 1
            continue

        # --- Create the WearMeasurement record ---
        meas = WearMeasurement(
            chainage_id=chainage_obj.id,
            measurement_date=measurement_date,
            wear_mm=round(wear_mm, 2),
            left_wear_0=round(left_0, 2) if left_0 is not None else None,
            left_wear_90=round(left_90, 2) if left_90 is not None else None,
            right_wear_0=round(right_0, 2) if right_0 is not None else None,
            right_wear_90=round(right_90, 2) if right_90 is not None else None,
            # NEL equipment only measures at 0 and 90 degrees
            left_wear_22_5=None,
            left_wear_45=None,
            left_wear_67_5=None,
            right_wear_22_5=None,
            right_wear_45=None,
            right_wear_67_5=None,
            source_file=filename,
        )
        result.measurements.append(meas)
        existing_pairs.add((chainage_obj.id, measurement_date))
        result.rows_accepted += 1

    # Set metadata on the result
    result.chainage_range = (int(min_ch) if min_ch != float("inf") else 0,
                              int(max_ch) if max_ch != float("-inf") else 0)
    result.sectors_found = sorted(sectors_seen)
    result.category_summary = cat_counts

    return result
